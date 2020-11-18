from django.shortcuts import render, HttpResponse, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from .models import RegForm
import math
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
import os
import smtplib
from email.message import EmailMessage


def home(request):
    return render(request, 'registration_ongc/home.html')


def register(request):
    if request.method == 'POST':
        name = request.POST['name']
        dob = request.POST['dob']
        gender = request.POST['gender']
        address = request.POST['address']
        mobile = request.POST['mobile']
        email = request.POST['mailadd']
        aadhar = request.POST['aadhar']
        fname = request.POST['father']
        foccupation = request.POST['occupation']
        institutename = request.POST['institute']
        course = request.POST['course']
        presentyear = request.POST['pyear']
        percentage12 = request.POST['std12']
        cgpa = request.POST['cgpa']
        photograph = request.FILES['photo']
        marksheet12 = request.FILES['marksheet12']
        lmarksheet = request.FILES['dmc']
        bletter = request.FILES['certificate']
        Designation = request.POST['designation']
        CPF = request.POST['CPF']
        Section = request.POST['section']
        Location = request.POST['location']
        PhoneNumber = request.POST['phoneNo']
        MobileNumber = request.POST['parMobile']
        details = RegForm(name=name, dob=dob, gender=gender, address=address, mobile=mobile, email=email, aadhar=aadhar, fname=fname, foccupation=foccupation, institutename=institutename, course=course, presentyear=presentyear, percentage12=percentage12,
                          cgpa=cgpa, photograph=photograph, marksheet12=marksheet12, lmarksheet=lmarksheet, bletter=bletter, Designation=Designation, CPF=CPF, Section=Section, Location=Location, PhoneNumber=PhoneNumber, MobileNumber=MobileNumber)
        details.save()
        print(name)
        print('we are using post request')
        all_xls_sheet()
        selected_xls_sheet()
        rejected_xls_sheet()
        selected_xls_sheet()
        messages.success(request, "Form submitted Successfully")
    return render(request, 'registration_ongc/registration.html')


def handleofficerLogin(request):
    if request.method == 'POST':
        officerusername = request.POST['officerusername']
        officerpass = request.POST['officerpass']
        user = authenticate(username=officerusername, password=officerpass)
        if user is not None:
            login(request, user)
            return redirect("officer")
        else:
            messages.error(request, "Invalid credentials")
            return redirect("/")
    return HttpResponse('404 - Not found')


def handleofficerLogout(request):

    logout(request)
    messages.success(request, "Successfully logout")
    return redirect("/")


@login_required(login_url='officerlogin')
def handelofficer(request):
    RegForms = RegForm.objects.all()
    n = len(RegForms)
    params = {'allforms': RegForms, 'range': range(n)}
    return render(request, 'registration_ongc/officer.html', params)


@login_required(login_url='officerlogin')
def deleterecent(request):
    if request.method == 'POST':
        id12 = request.POST['dvalue']
        d = RegForm.objects.get(sno=id12)
        d.counter = 1
        d.save()
        selected_xls_sheet()
        all_xls_sheet()
        rejected_xls_sheet()
        selected_xls_sheet()
        return redirect("officer")


@login_required(login_url='officerlogin')
def acceptrecent(request):
    if request.method == 'POST':
        id12 = request.POST['dvalue1']
        d = RegForm.objects.get(sno=id12)
        d.counter = 2
        d.save()
        selected_xls_sheet()
        all_xls_sheet()
        rejected_xls_sheet()
        selected_xls_sheet()
        return redirect("officer")


def selected_xls_sheet():
    RegForms = RegForm.objects.all()
    total_no = len(RegForms)
    c = 2
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 50
    ws1.column_dimensions['H'].width = 20
    ws1.column_dimensions['I'].width = 20
    ws1.column_dimensions['J'].width = 25
    ws1.column_dimensions['K'].width = 25
    ws1.column_dimensions['L'].width = 15
    ws1.column_dimensions['M'].width = 15
    ws1.column_dimensions['N'].width = 15
    ws1.column_dimensions['O'].width = 20
    ws1.column_dimensions['P'].width = 12
    ws1.column_dimensions['Q'].width = 12
    ws1.column_dimensions['R'].width = 12
    ws1.column_dimensions['S'].width = 20
    ws1.column_dimensions['T'].width = 20
    (ws1.cell(row=1, column=1, value="Name")).font = Font(bold=True)
    (ws1.cell(row=1, column=2, value="DOB")).font = Font(bold=True)
    (ws1.cell(row=1, column=3, value="Gender")).font = Font(bold=True)
    (ws1.cell(row=1, column=4, value="E-mail")).font = Font(bold=True)
    (ws1.cell(row=1, column=5, value="Adhaar Number")).font = Font(bold=True)
    (ws1.cell(row=1, column=6, value="Mobile")).font = Font(bold=True)
    (ws1.cell(row=1, column=7, value="Address")).font = Font(bold=True)
    (ws1.cell(row=1, column=8, value="Father/Mother Name")).font = Font(bold=True)
    (ws1.cell(row=1, column=9, value="Occupation")).font = Font(bold=True)
    (ws1.cell(row=1, column=10, value="Institute Name")).font = Font(bold=True)
    (ws1.cell(row=1, column=11, value="Course")).font = Font(bold=True)
    (ws1.cell(row=1, column=12, value="Present Year")).font = Font(bold=True)
    (ws1.cell(row=1, column=13, value="12th Percentage")).font = Font(bold=True)
    (ws1.cell(row=1, column=14, value="CGPA")).font = Font(bold=True)
    (ws1.cell(row=1, column=15, value="Designation")
     ).font = Font(color="FF0000", bold=True)
    (ws1.cell(row=1, column=16, value="CPF")
     ).font = Font(color="FF0000", bold=True)
    (ws1.cell(row=1, column=17, value="Section")
     ).font = Font(color="FF0000", bold=True)
    (ws1.cell(row=1, column=18, value="Location")
     ).font = Font(color="FF0000", bold=True)
    (ws1.cell(row=1, column=19, value="Phone Number")
     ).font = Font(color="FF0000", bold=True)
    (ws1.cell(row=1, column=20, value="Mobile Number")
     ).font = Font(color="FF0000", bold=True)

    for i in range(total_no):
        name = RegForms[i].name
        dob = str(RegForms[i].dob)
        gender = RegForms[i].gender
        address = RegForms[i].address
        mobile = RegForms[i].mobile
        email = RegForms[i].email
        aadhar = str(RegForms[i].aadhar)
        fname = RegForms[i].fname
        foccupation = RegForms[i].foccupation
        institutename = RegForms[i].institutename
        course = RegForms[i].course
        presentyear = RegForms[i].presentyear
        percentage12 = RegForms[i].percentage12
        cgpa = RegForms[i].cgpa
        Designation = RegForms[i].Designation
        CPF = RegForms[i].CPF
        Section = RegForms[i].Section
        Location = RegForms[i].Location
        PhoneNumber = RegForms[i].PhoneNumber
        MobileNumber = RegForms[i].MobileNumber
        if RegForms[i].counter == 2:
            ws1.cell(row=c, column=1, value=name)
            ws1.cell(row=c, column=2, value=dob)
            ws1.cell(row=c, column=3, value=gender)
            ws1.cell(row=c, column=4, value=email)
            ws1.cell(row=c, column=5, value=aadhar)
            ws1.cell(row=c, column=6, value=mobile)
            ws1.cell(row=c, column=7, value=address)
            ws1.cell(row=c, column=8, value=fname)
            ws1.cell(row=c, column=9, value=foccupation)
            ws1.cell(row=c, column=10, value=institutename)
            ws1.cell(row=c, column=11, value=course)
            ws1.cell(row=c, column=12, value=presentyear)
            ws1.cell(row=c, column=13, value=percentage12)
            ws1.cell(row=c, column=14, value=cgpa)
            ws1.cell(row=c, column=15, value=Designation)
            ws1.cell(row=c, column=16, value=CPF)
            ws1.cell(row=c, column=17, value=Section)
            ws1.cell(row=c, column=18, value=Location)
            ws1.cell(row=c, column=19, value=PhoneNumber)
            ws1.cell(row=c, column=20, value=MobileNumber)
            c = c+1
        wb1.save("media/xl_sheets1/selected_one_xl_file.xlsx")


def rejected_xls_sheet():
    RegForms = RegForm.objects.all()
    total_no = len(RegForms)
    c = 2
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 12
    ws.column_dimensions['Q'].width = 12
    ws.column_dimensions['R'].width = 12
    ws.column_dimensions['S'].width = 20
    ws.column_dimensions['T'].width = 20
    (ws.cell(row=1, column=1, value="Name")).font = Font(bold=True)
    (ws.cell(row=1, column=2, value="DOB")).font = Font(bold=True)
    (ws.cell(row=1, column=3, value="Gender")).font = Font(bold=True)
    (ws.cell(row=1, column=4, value="E-mail")).font = Font(bold=True)
    (ws.cell(row=1, column=5, value="Adhaar Number")).font = Font(bold=True)
    (ws.cell(row=1, column=6, value="Mobile")).font = Font(bold=True)
    (ws.cell(row=1, column=7, value="Address")).font = Font(bold=True)
    (ws.cell(row=1, column=8, value="Father/Mother Name")).font = Font(bold=True)
    (ws.cell(row=1, column=9, value="Occupation")).font = Font(bold=True)
    (ws.cell(row=1, column=10, value="Institute Name")).font = Font(bold=True)
    (ws.cell(row=1, column=11, value="Course")).font = Font(bold=True)
    (ws.cell(row=1, column=12, value="Present Year")).font = Font(bold=True)
    (ws.cell(row=1, column=13, value="12th Percentage")).font = Font(bold=True)
    (ws.cell(row=1, column=14, value="CGPA")).font = Font(bold=True)
    (ws.cell(row=1, column=15, value="Designation")
     ).font = Font(color="FF0000", bold=True)
    (ws.cell(row=1, column=16, value="CPF")
     ).font = Font(color="FF0000", bold=True)
    (ws.cell(row=1, column=17, value="Section")
     ).font = Font(color="FF0000", bold=True)
    (ws.cell(row=1, column=18, value="Location")
     ).font = Font(color="FF0000", bold=True)
    (ws.cell(row=1, column=19, value="Phone Number")
     ).font = Font(color="FF0000", bold=True)
    (ws.cell(row=1, column=20, value="Mobile Number")
     ).font = Font(color="FF0000", bold=True)
    for i in range(total_no):
        name = RegForms[i].name
        dob = str(RegForms[i].dob)
        gender = RegForms[i].gender
        address = RegForms[i].address
        mobile = RegForms[i].mobile
        email = RegForms[i].email
        aadhar = str(RegForms[i].aadhar)
        fname = RegForms[i].fname
        foccupation = RegForms[i].foccupation
        institutename = RegForms[i].institutename
        course = RegForms[i].course
        presentyear = RegForms[i].presentyear
        percentage12 = RegForms[i].percentage12
        cgpa = RegForms[i].cgpa
        Designation = RegForms[i].Designation
        CPF = RegForms[i].CPF
        Section = RegForms[i].Section
        Location = RegForms[i].Location
        PhoneNumber = RegForms[i].PhoneNumber
        MobileNumber = RegForms[i].MobileNumber
        if RegForms[i].counter == 1:
            ws.cell(row=c, column=1, value=name)
            ws.cell(row=c, column=2, value=dob)
            ws.cell(row=c, column=3, value=gender)
            ws.cell(row=c, column=4, value=email)
            ws.cell(row=c, column=5, value=aadhar)
            ws.cell(row=c, column=6, value=mobile)
            ws.cell(row=c, column=7, value=address)
            ws.cell(row=c, column=8, value=fname)
            ws.cell(row=c, column=9, value=foccupation)
            ws.cell(row=c, column=10, value=institutename)
            ws.cell(row=c, column=11, value=course)
            ws.cell(row=c, column=12, value=presentyear)
            ws.cell(row=c, column=13, value=percentage12)
            ws.cell(row=c, column=14, value=cgpa)
            ws.cell(row=c, column=15, value=Designation)
            ws.cell(row=c, column=16, value=CPF)
            ws.cell(row=c, column=17, value=Section)
            ws.cell(row=c, column=18, value=Location)
            ws.cell(row=c, column=19, value=PhoneNumber)
            ws.cell(row=c, column=20, value=MobileNumber)
            c = c+1
    wb.save("media/xl_sheets1/rejected_one_xl_file.xlsx")


def all_xls_sheet():
    RegForms = RegForm.objects.all()
    total_no = len(RegForms)
    c=2
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 25
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 12
    ws.column_dimensions['R'].width = 12
    ws.column_dimensions['S'].width = 12
    ws.column_dimensions['T'].width = 20
    ws.column_dimensions['U'].width = 20
    (ws.cell(row=1, column=1, value="Name")).font = Font(bold=True)
    (ws.cell(row=1, column=2, value="Status")).font = Font(bold=True)
    (ws.cell(row=1, column=3, value="DOB")).font = Font(bold=True)
    (ws.cell(row=1, column=4, value="Gender")).font = Font(bold=True)
    (ws.cell(row=1, column=5, value="E-mail")).font = Font(bold=True)
    (ws.cell(row=1, column=6, value="Adhaar Number")).font = Font(bold=True)
    (ws.cell(row=1, column=7, value="Mobile")).font = Font(bold=True)
    (ws.cell(row=1, column=8, value="Address")).font = Font(bold=True)
    (ws.cell(row=1, column=9, value="Father/Mother Name")).font = Font(bold=True)
    (ws.cell(row=1, column=10, value="Occupation")).font = Font(bold=True)
    (ws.cell(row=1, column=11, value="Institute Name")).font = Font(bold=True)
    (ws.cell(row=1, column=12, value="Course")).font = Font(bold=True)
    (ws.cell(row=1, column=13, value="Present Year")).font = Font(bold=True)
    (ws.cell(row=1, column=14, value="12th Percentage")).font = Font(bold=True)
    (ws.cell(row=1, column=15, value="CGPA")).font = Font(bold=True)
    (ws.cell(row=1, column=16, value="Designation")
     ).font = Font(color="FF0000", bold=True)
    (ws.cell(row=1, column=17, value="CPF")
     ).font = Font(color="FF0000", bold=True)
    (ws.cell(row=1, column=18, value="Section")
     ).font = Font(color="FF0000", bold=True)
    (ws.cell(row=1, column=19, value="Location")
     ).font = Font(color="FF0000", bold=True)
    (ws.cell(row=1, column=20, value="Phone Number")
     ).font = Font(color="FF0000", bold=True)
    (ws.cell(row=1, column=21, value="Mobile Number")
     ).font = Font(color="FF0000", bold=True)
    for i in range(total_no):
        if RegForms[i].counter == 1 or RegForms[i].counter == 2:
            if RegForms[i].counter == 1:
                status = "ACCEPTED"
            else:
                status = "REJECTED"
            name = RegForms[i].name
            dob = str(RegForms[i].dob)
            gender = RegForms[i].gender
            address = RegForms[i].address
            mobile = RegForms[i].mobile
            email = RegForms[i].email
            aadhar = str(RegForms[i].aadhar)
            fname = RegForms[i].fname
            foccupation = RegForms[i].foccupation
            institutename = RegForms[i].institutename
            course = RegForms[i].course
            presentyear = RegForms[i].presentyear
            percentage12 = RegForms[i].percentage12
            cgpa = RegForms[i].cgpa
            Designation = RegForms[i].Designation
            CPF = RegForms[i].CPF
            Section = RegForms[i].Section
            Location = RegForms[i].Location
            PhoneNumber = RegForms[i].PhoneNumber
            MobileNumber = RegForms[i].MobileNumber
            ws.cell(row=c, column=1, value=name)
            ws.cell(row=c, column=2, value=status)
            ws.cell(row=c, column=3, value=dob)
            ws.cell(row=c, column=4, value=gender)
            ws.cell(row=c, column=5, value=email)
            ws.cell(row=c, column=6, value=aadhar)
            ws.cell(row=c, column=7, value=mobile)
            ws.cell(row=c, column=8, value=address)
            ws.cell(row=c, column=9, value=fname)
            ws.cell(row=c, column=10, value=foccupation)
            ws.cell(row=c, column=11, value=institutename)
            ws.cell(row=c, column=12, value=course)
            ws.cell(row=c, column=13, value=presentyear)
            ws.cell(row=c, column=14, value=percentage12)
            ws.cell(row=c, column=15, value=cgpa)
            ws.cell(row=c, column=16, value=Designation)
            ws.cell(row=c, column=17, value=CPF)
            ws.cell(row=c, column=18, value=Section)
            ws.cell(row=c, column=19, value=Location)
            ws.cell(row=c, column=20, value=PhoneNumber)
            ws.cell(row=c, column=21, value=MobileNumber)
            c=c+1
        # Save the file
    wb.save("media/xl_sheets1/all_xl_file1.xlsx")



# $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
#                           MAIL
# $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$



def send_email(to,name,mobile):
    name = name.upper()
    mobile = mobile
    message = EmailMessage()
    message['subject'] = "summer training-2020"
    message['from'] = "user email"
    message['to'] = to
    message.set_content("hello ")
    html_message = open("registration_ongc/templates/registration_ongc/mail.html",encoding="utf8").read().replace("Taat",name+" - "+mobile)
    message.add_alternative(html_message, subtype='html')
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login("username", "password")
        smtp.send_message(message)

@login_required(login_url='officerlogin')
def sendmails(request):
    RegForms = RegForm.objects.all()
    total_no = len(RegForms)
    for i in range(total_no):
        if RegForms[i].counter == 2:
            name = RegForms[i].name
            mobile = RegForms[i].mobile
            email = RegForms[i].email 
            send_email(email,name,mobile)
    messages.success(request, "Mail sent")
    return redirect("officer")